# pylint: disable=E0611
# pylint: disable=E1101

from PyQt5.QtGui import QBrush, QImage, QPixmap, QPainter, QPen, QCursor,QPolygonF, QWheelEvent,QColor, QMouseEvent,\
     QStandardItemModel, QStandardItem, QIcon, QWheelEvent
from PyQt5.QtWidgets import QMainWindow, QWidget, QGraphicsScene, QGraphicsView, QAction, QActionGroup,\
     QPushButton, QGridLayout, QApplication, QVBoxLayout, QHBoxLayout, QGraphicsEllipseItem, QGraphicsItem,\
     QGraphicsItemGroup,QGraphicsSceneMouseEvent, QListView, QSplitter, QFrame, QSizePolicy, QTreeView,\
     QHeaderView, QCheckBox, QComboBox, QFileDialog, QTabWidget, QTableWidget, QSpinBox, QLabel, QTableWidgetItem,\
     QColorDialog, QProgressDialog, QTreeWidget, QTreeWidgetItem, QMessageBox,QLineEdit, QDoubleSpinBox
from PyQt5.QtCore import Qt, QPoint, QLineF,QPointF, QEvent, QPersistentModelIndex, QModelIndex

from DrawObjects import GraphicsCircleItem,GraphicsLineItem,GraphicsPolylineItem,\
     GraphicsRectItem, OneCalcCircle, RecPolylineItem
from ObjectsMenu import DownloadDelegate, Save_Widget, MenuData, MyValidator

from main_calculate import run_area_calc, point_calc, setTypes, cuda_available, receivers_calc


from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from matplotlib.colors import LogNorm, Normalize
from matplotlib.collections import LineCollection
from matplotlib.cm import jet
import matplotlib.pyplot as plt
import matplotlib.ticker
import matplotlib
from matplotlib.ticker import FuncFormatter
import matplotlib.font_manager as fm
arial_font = fm.FontProperties(fname = "Fonts/arial.ttf")

import numpy as np
import scipy.interpolate
from functools import partial

from openpyxl import load_workbook, Workbook

import dxf

import sys
import os

import json
import pickle

import inspect

from time import time


class TreeWidgetItem(QTreeWidgetItem):
    def __hash__(self):
        own_hash = hash(str(self))
        return own_hash



class GraphicsView(QGraphicsView):
    def __init__(self,SoursesObjectsDict, AreasObjectsDict, ReceiverObjectsDict, ReturnMousePos = None):
        super().__init__()
        self.SoursesObjectsDict = SoursesObjectsDict
        self.AreasObjectsDict = AreasObjectsDict
        self.ReceiverObjectsDict = ReceiverObjectsDict
        self.ReturnMousePos = ReturnMousePos
        self.current_scale = 1
        self.setViewportUpdateMode(QGraphicsView.FullViewportUpdate)
        #self.setCacheMode(QGraphicsView.CacheNone)

        self.grid_step()

    @staticmethod
    def digit(x):
        a = len(str(int(x)))
        if a>2:
            a -= 2
            return int(x)/10**a, a
        else:
            return int(x), 0

    def wheelEvent(self, WheelEvent):
        """ Событие скрола """
        #super(Screen, self).wheelEvent(QWheelEvent)
        wheelcounter = WheelEvent.angleDelta()
        

        if wheelcounter.y()==120:
            self.scale(1.25,1.25)
            self.current_scale*=1.25
            for obj in self.SoursesObjectsDict.values():
                if obj.type_link == "object":
                    obj.graphic_item.hndl*=0.8
                    obj.graphic_item.updateHandlesPos()                   

            for obj in self.AreasObjectsDict.values():
                if obj.type_link == "object":
                    obj.graphic_item.hndl*=0.8
                    obj.graphic_item.updateHandlesPos() 

            for obj in self.ReceiverObjectsDict.values():
                if obj.type_link == "object":
                    obj.graphic_item.hndl*=0.8
                    obj.graphic_item.updateHandlesPos() 
                    

        elif wheelcounter.y()==-120:
            self.scale(0.8,0.8)
            self.current_scale*=0.8
            for obj in self.SoursesObjectsDict.values():
                if obj.type_link == "object":
                    obj.graphic_item.hndl*=1.25
                    obj.graphic_item.updateHandlesPos()                   

            for obj in self.AreasObjectsDict.values():
                if obj.type_link == "object":
                    obj.graphic_item.hndl*=1.25
                    obj.graphic_item.updateHandlesPos()

            for obj in self.ReceiverObjectsDict.values():
                if obj.type_link == "object":
                    obj.graphic_item.hndl*=1.25
                    obj.graphic_item.updateHandlesPos()


    def scale(self,sx,sy):
        QGraphicsView.scale(self,sx,sy)
        self.grid_step()       


    def grid_step(self):
        rect = self.mapToScene(self.rect()).boundingRect()
        a, b = rect.bottom(), rect.top()
        c, d = rect.left(), rect.right()

        if a <= b: startY,stopY = int(a), int(b)
        else: startY,stopY = int(b), int(a)
        if c <= d: startX,stopX = int(c), int(d)
        else: startX,stopX = int(d), int(c)

        w,h = self.size().width(), self.size().height()
        #print(w,h)
        #print(stopX-startX,stopY-startY)
        #step = (min(stopX-startX,stopY-startY)*min(w,h))/(7000)
        step = (min(stopX-startX,stopY-startY)*30)/min(w,h)
        d, st = self.digit(step)
        r = [1,5,10,25,50]
        self.step = min(map(lambda x: (x,x-d) , r),key=lambda x: abs(x[1]))[0] *10**st
        self.line_width = min(w,h)/self.current_scale*0.001
        #print(self.step)


    def mousePressEvent(self, mouseEvent):
        if mouseEvent.button() == 4: # номер кнопки колёсика
            self.setDragMode(QGraphicsView.ScrollHandDrag)
            self.setInteractive(False)

            click = QMouseEvent(QEvent.GraphicsSceneMousePress,mouseEvent.pos(), Qt.LeftButton,\
                        Qt.LeftButton, Qt.NoModifier)
            self.mousePressEvent(click)

        super().mousePressEvent(mouseEvent)

    def mouseMoveEvent(self, mouseEvent):
        if self.ReturnMousePos is not None:
            xy = self.mapToScene(mouseEvent.pos())
            self.ReturnMousePos(round(xy.x()/1000,3), round(-xy.y()/1000,3))
        super().mouseMoveEvent(mouseEvent)

    def mouseReleaseEvent(self, mouseEvent):
        super().mouseReleaseEvent(mouseEvent)
        self.setDragMode(QGraphicsView.NoDrag)
        self.setInteractive(True)


    def drawBackground(self,painter,rect, n="a"):
        step = self.step
        if n == "a":
            pen = QPen(QColor(0, 0, 0, 190))
        else:
            pen = QPen(QColor(255, 0, 0, 190))
        pen.setWidthF(self.line_width)
        painter.setPen(pen)
        a, b = rect.bottom(), rect.top()
        c, d = rect.left(), rect.right()

        if a <= b: startY,stopY = int((a//step)*step), int(b)
        else: startY,stopY = int((b//step)*step), int(a)
        if c <= d: startX,stopX = int((c//step)*step), int(d)
        else: startX,stopX = int((d//step)*step), int(c)
         
        for y in range(startY,stopY,step):
            painter.drawLine(c, y, d, y)
        for x in range(startX,stopX,step):
            painter.drawLine(x, a, x, b)


        


class GraphicsScene(QGraphicsScene):
    def __init__(self, parent=None):
        QGraphicsScene.__init__(self, parent)

    def dragable(self,trig):
        for i in self.items():
            i.dragable(trig)

        
# subclass
class CheckableComboBox(QComboBox):
    # once there is a checkState set, it is rendered
    # here we assume default Unchecked
    def __init__(self,parent=None):
        super(CheckableComboBox, self).__init__(parent)
        self.title_text = "Слои"
        self.model().itemChanged.connect(self.itemChecked)
        self.setEditable(True)
        self.displayedText = self.lineEdit()
        self.displayedText.setText(self.title_text)
        self.displayedText.setReadOnly(True)
        self.isCheckFunc = False

    def addItem(self, item):
        super(CheckableComboBox, self).addItem(item)
        item = self.model().item(self.count()-1,0)
        item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
        item.setCheckState(Qt.Unchecked)

    def getItem(self,text):
        try:
            item = self.model().item(self.findText(text),0)
            return item
        except Exception as ex:
            print(ex)
            return None
    def setItemState(self, text, state):
        try:
            item = self.model().item(self.findText(text),0)
            item.setCheckState(Qt.Checked if state else Qt.Unchecked)
        except Exception as ex:
            print(ex)
            

    def paintEvent(self,ev):
        self.displayedText.setText(self.title_text)
        super().paintEvent(ev)

    def setCheckEvent(self,func):
        self.Func = func


    def itemChecked(self, item):
        if self.isCheckFunc:
            self.Func(item.text(),True if item.checkState()==Qt.Checked else False)

    def ShowNewObj(self,text):
        try:
            item = self.model().item(self.findText(text),0)
            self.Func(item.text(),True if item.checkState()==Qt.Checked else False)
        except Exception as ex:
            print(ex)

    def removeItem(self, text):
        self.model().removeRow(self.findText(text))


    """ def handleItemPressed(self, index):
        item = self.model().itemFromIndex(index)
        if item.checkState() == Qt.Checked:
            item.setCheckState(Qt.Unchecked)
        else:
            item.setCheckState(Qt.Checked) """



class Atributs:
    def __init__(self,type_link = "layer",
        type_object = None, graphic_item = None,
        movable = True, state = False):

        self.type_link = type_link
        self.type_object = type_object
        self.graphic_item = graphic_item
        self.movable = movable
        self.state = state

        
        
    

class Screen(QMainWindow):
    def __init__(self):
        super().__init__()

        desktop = QApplication.desktop()
        wd = desktop.width()
        hg = desktop.height()
        ww = 1000
        wh = 500
        if ww>wd: ww = int(0.7*wd)
        if wh>hg: wh = int(0.7*hg)
        x = (wd-ww)//2
        y = (hg-wh)//2
        self.setGeometry(x, y, ww, wh)

        try:
            self.path_home = os.path.expanduser("~\\Desktop\\")
        except Exception:
            self.path_home = ""

        for curren_dir in ["Cashe_files"]:
            if os.path.exists(curren_dir):
                if os.path.isdir(curren_dir):
                    print(curren_dir+" is here")
                else:
                    try:
                        os.mkdir(curren_dir)
                    except OSError:
                        print ("Error generate dir "+curren_dir)
            else:
                try:
                    os.mkdir(curren_dir)
                except OSError:
                    print ("Error generate dir "+curren_dir)

        self.setWindowTitle('MFC')
        self.setWindowIcon(QIcon("images/icon.png"))

        LeftPanelFrame = QFrame() 
        LeftPanelFrame.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        #LeftPanelFrame.setMinimumSize(QSize(0, 100))

        # Treewidgets for all user objects
        self.SoursesObjectsTree = QTreeWidget()
        self.AreasObjectsTree = QTreeWidget()
        self.ReceiverObjectsTree = QTreeWidget()

        """ self.SoursesObjectsTree.itemClicked.connect(lambda a,b:print("Click",a,b))
        self.SoursesObjectsTree.itemActivated.connect(lambda a,b:print("Activated",a,b))
        self.SoursesObjectsTree.itemChanged.connect(lambda a,b:print("Changed",a,b))
        self.SoursesObjectsTree.itemPressed.connect(lambda a,b:print("Pressed",a,b))
        self.SoursesObjectsTree.itemDoubleClicked.connect(lambda a,b:print("DoubleClicked",a,b)) """

        self.pcalc = False

        self.SoursesObjectsTree.setColumnCount(2)
        self.SoursesObjectsTree.setHeaderLabels(["","Источники"])
        self.SoursesObjectsTree.setColumnWidth(0, 60) 

        self.AreasObjectsTree.setColumnCount(2)
        self.AreasObjectsTree.setHeaderLabels(["","Области расчета"])
        self.AreasObjectsTree.setColumnWidth(0, 60) 

        self.ReceiverObjectsTree.setColumnCount(2)
        self.ReceiverObjectsTree.setHeaderLabels(["","Приемники"])
        self.ReceiverObjectsTree.setColumnWidth(0, 60) 

        self.SoursesObjectsTree.itemChanged.connect(lambda a,b:self.ChangeCheckBox(a,b,"sourses"))
        self.SoursesObjectsTree.itemActivated.connect(lambda a,b:self.OpenObjMenu(a,b,"sourses"))

        self.AreasObjectsTree.itemChanged.connect(lambda a,b:self.ChangeCheckBox(a,b,"areas"))
        self.AreasObjectsTree.itemActivated.connect(lambda a,b:self.OpenObjMenu(a,b,"areas"))

        self.ReceiverObjectsTree.itemChanged.connect(lambda a,b:self.ChangeCheckBox(a,b,"receivers"))
        self.ReceiverObjectsTree.itemActivated.connect(lambda a,b:self.OpenObjMenu(a,b,"receivers"))

        #self.SoursesObjectsTree.setEditTriggers(QTreeWidget.DoubleClicked | QTreeWidget.SelectedClicked | QTreeWidget.EditKeyPressed)

        # Dictionaries for all user objects
        self.SoursesObjectsDict = {}
        self.AreasObjectsDict = {}
        self.ReceiverObjectsDict = {}

        self.SoursesObjectsChildren = {}
        self.AreasObjectsChildren = {}
        self.ReceiverObjectsChildren = {}
 
        ListObjectsHeader = QHeaderView(Qt.Horizontal)

        da = QLabel("𝑑𝛼, \u00B0:")
        da.setStyleSheet("font: 10pt;")
        dl = QLabel("𝑑𝑙, м:")
        dl.setStyleSheet("font: 10pt;")

        self.da = QDoubleSpinBox()
        self.da.setRange(0.001,360)
        self.da.setSingleStep(1)
        self.da.setValue(1)
        #self.da.editingFinished.connect()

        self.dl = QDoubleSpinBox()
        self.dl.setRange(0.001,99)
        self.dl.setSingleStep(0.01)
        self.dl.setValue(0.01)
        #self.dl.editingFinished.connect()

        calc_setings = QHBoxLayout()
        calc_setings.addWidget(da)
        calc_setings.addWidget(self.da)
        calc_setings.addWidget(dl)
        calc_setings.addWidget(self.dl)

        # Left pannel with treewidges
        BoxLayout2 = QHBoxLayout()
        BoxLayout2_1 = QVBoxLayout()
        BoxLayout2_1.addWidget(self.SoursesObjectsTree)
        BoxLayout2_1.addLayout(calc_setings)
        BoxLayout2_1.addWidget(self.AreasObjectsTree)
        BoxLayout2_1.addWidget(self.ReceiverObjectsTree)

        BoxLayout2.addLayout(BoxLayout2_1)
        LeftPanelFrame.setLayout(BoxLayout2) 

        SceneFrame = QFrame() 
        SceneFrame.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding) 
        #SceneFrame.setMinimumSize(QSize(0, 100))

        self.status_string = self.statusBar() #.showMessage('Ready')

        pic = QPixmap("images/%s.png" % ("gpu" if cuda_available() else "cpu"))
        scaled=pic.scaled(31, 20, Qt.IgnoreAspectRatio, Qt.FastTransformation)
        label = QLabel()
        label.setPixmap(scaled)
        Phbox = QHBoxLayout(self)
        Phbox.addWidget(label)
        Phbox.setContentsMargins(0,0,4,4)

        Pct = QWidget()
        #Pct.setStyleSheet("background-color: red")
        #Pct.setStyleSheet ("border: 1px solid red")
        Pct.setLayout(Phbox)
 
        self.status_string.addPermanentWidget(Pct)

        cords_string = lambda x,y: self.status_string.showMessage(f"x: {x}, y: {y}")

        self.scene = GraphicsScene() #QGraphicsScene()
        self.view = GraphicsView(self.SoursesObjectsDict, self.AreasObjectsDict, self.ReceiverObjectsDict, cords_string)
        self.view.setMouseTracking(True)
        self.view.setAlignment( Qt.AlignLeft | Qt.AlignTop )
        self.view.setScene(self.scene)
        self.view.current_scale = 1

        self.tab = QTabWidget()
        self.tab.addTab(self.view,"Модель")
        self.fig = plt.figure(dpi=75)
        self.Canv = FigureCanvas(self.fig)

        self.rbf = None
        self.last_fig = None # Mfield, Irec

        self.Canv.setFocusPolicy(Qt.ClickFocus)
        self.Canv.setFocus()

        #self.fig.canvas.mpl_connect('button_press_event', self.on_mouse_move)
        self.fig.canvas.mpl_connect('motion_notify_event', self.on_mouse_move)
        


        self.table_contour = QTableWidget()
        self.table_contour.setColumnCount(3)
        #self.table_contour.setRowCount(2)
        self.table_contour.setColumnWidth(0,30)
        self.table_contour.setColumnWidth(1,70)
        self.table_contour.setColumnWidth(2,70)
        self.table_contour.setHorizontalHeaderLabels(["","H, А/м","Цвет"])
        self.table_contour.setMaximumWidth(200)
        self.table_contour.setItemDelegate(DownloadDelegate("contour",self))
        self.table_contour.cellClicked[int,int].connect(self.SetColorTContour) #partial(self.OpenFig,2)
        self.table_contour.cellChanged.connect(self.SetCountur)
        self.d_contur = {}
        self.make_contur = None
        self.old_rows = 0

        self.tbc_rows = QSpinBox()
        self.tbc_rows.setRange(0,999)
        self.tbc_rows.setSingleStep(1)
        self.tbc_rows.setValue(0)
        self.tbc_rows.editingFinished.connect(self.ResizeTablesContour)

        BoxLayout3 = QVBoxLayout()
        BoxLayout3.addWidget(QLabel("Линии уровня"))
        BoxLayout3.addWidget(self.tbc_rows)
        BoxLayout3.addWidget(self.table_contour)

        BoxLayout4 = QHBoxLayout()
        BoxLayout4.addWidget(self.Canv,1)
        BoxLayout4.addLayout(BoxLayout3,0)
        Rez_widget = QWidget()
        Rez_widget.setLayout(BoxLayout4)
        self.tab.addTab(Rez_widget,"Результаты")


        BoxLayout1 = QHBoxLayout()
        BoxLayout1.addWidget(self.tab)
        SceneFrame.setLayout(BoxLayout1) 

        
 
        menubar = self.menuBar()
        fileMenu = menubar.addMenu('&Файл')

        OpenDXFAction = QAction('Открыть .dxf', self)
        OpenDXFAction.setShortcut('Ctrl+O')
        OpenDXFAction.triggered.connect(self.load_dxf_file) 
        fileMenu.addAction(OpenDXFAction)

        OpenSelfAction = QAction('Открыть', self)
        #OpenSelfAction.setShortcut('Ctrl+O')
        OpenSelfAction.triggered.connect(self.LoadCalcData) 
        fileMenu.addAction(OpenSelfAction)

        SaveObjAction = QAction('Сохранить', self)
        SaveObjAction.setShortcut('Ctrl+S')
        SaveObjAction.triggered.connect(self.SaveCalcData) 
        fileMenu.addAction(SaveObjAction)


        #CorrectMenu = menubar.addMenu('&Правка')

 
        SourceMenu = menubar.addMenu('&Источники')

        NewLayerAction_1 = QAction('Добавить слой', self)
        #NewLayerAction.setShortcut('Ctrl+R')
        NewLayerAction_1.triggered.connect(lambda:self.AddObj("sourses","layer")) 
        SourceMenu.addAction(NewLayerAction_1)

        NewReactorAction = QAction('Добавить реактор', self)
        NewReactorAction.setShortcut('Ctrl+R')
        NewReactorAction.triggered.connect(lambda:self.AddObj("sourses","reactor"))
        SourceMenu.addAction(NewReactorAction)

        NewWireAction = QAction('Добавить шину', self)
        NewWireAction.setShortcut('Ctrl+W')
        NewWireAction.triggered.connect(lambda:self.AddObj("sourses","conductor"))
        SourceMenu.addAction(NewWireAction)

        DelObjectAction = QAction('Удалить', self)
        DelObjectAction.setShortcut('Ctrl+D')
        DelObjectAction.triggered.connect(lambda:self.DelObj("sourses")) 
        SourceMenu.addAction(DelObjectAction)


        AreaMenu = menubar.addMenu('&Области')

        NewLayerAction_2 = QAction('Добавить слой', self)
        #NewLayerAction_2.setShortcut('Ctrl+R')
        NewLayerAction_2.triggered.connect(lambda: self.AddObj("areas","layer")) 
        AreaMenu.addAction(NewLayerAction_2)

        NewHorizCalcAction = QAction('Горизонтальная область расчёта', self)
        NewHorizCalcAction.setShortcut('Ctrl+H')
        NewHorizCalcAction.triggered.connect(lambda: self.AddObj("areas","horizontal_area")) 
        AreaMenu.addAction(NewHorizCalcAction)

        NewVertCalcAction = QAction('Вертикальная область расчёта', self)
        NewVertCalcAction.setShortcut('Ctrl+G')
        NewVertCalcAction.triggered.connect(lambda: self.AddObj("areas","vertical_area")) 
        AreaMenu.addAction(NewVertCalcAction)

        NewPointCalcAction = QAction('Расчётная точка в пространстве', self)
        NewPointCalcAction.setShortcut('Ctrl+J')
        NewPointCalcAction.triggered.connect(lambda: self.AddObj("areas","one_point")) 
        AreaMenu.addAction(NewPointCalcAction)

        DelCalcAction = QAction('Удалить', self)
        DelCalcAction.setShortcut('Ctrl+K')
        DelCalcAction.triggered.connect(lambda:self.DelObj("areas")) 
        AreaMenu.addAction(DelCalcAction)


        ReceiverMenu = menubar.addMenu('&Приёмники')

        NewLayerAction_3 = QAction('Добавить слой', self)
        #NewLayerAction_3.setShortcut('Ctrl+R')
        NewLayerAction_3.triggered.connect(lambda:self.AddObj("receivers","layer")) 
        ReceiverMenu.addAction(NewLayerAction_3)

        NewRecconductorAction = QAction('Добавить проводник', self)
        #NewRecconductorAction.setShortcut('Ctrl+R')
        NewRecconductorAction.triggered.connect(lambda:self.AddObj("receivers","recconductor"))
        ReceiverMenu.addAction(NewRecconductorAction)

        DelObjectAction = QAction('Удалить', self)
        #DelObjectAction.setShortcut('Ctrl+D')
        DelObjectAction.triggered.connect(lambda:self.DelObj("receivers")) 
        ReceiverMenu.addAction(DelObjectAction)


        ResultMenu = menubar.addMenu('&Результаты')

        SaveInDXFAction = QAction('Сохранить в dxf', self)
        #SaveInDXFAction.setShortcut('Ctrl+R')
        SaveInDXFAction.triggered.connect(self.SaveInDXFstart) 
        ResultMenu.addAction(SaveInDXFAction)

        SavePlotAction = QAction('Сохранить график', self)
        #SaveInDXFAction.setShortcut('Ctrl+R')
        SavePlotAction.triggered.connect(self.SavePlot) 
        ResultMenu.addAction(SavePlotAction)

        RunCalcAction = QAction('Расчёт магнитного поля', self)
        #RunCalcAction.setShortcut('Ctrl+R')
        RunCalcAction.triggered.connect(self.Run_area) 
        ResultMenu.addAction(RunCalcAction)

        RecCurCalcAction = QAction('Расчёт тока в приёмниках', self)
        #RecCurCalcAction.setShortcut('Ctrl+R')
        RecCurCalcAction.triggered.connect(self.RecCalc) 
        ResultMenu.addAction(RecCurCalcAction)

        RecEcxSvAction = QAction('Соранить приемники в Excel', self)
        #RecCurCalcAction.setShortcut('Ctrl+R')
        RecEcxSvAction.triggered.connect(self.SaveExcel) 
        ResultMenu.addAction(RecEcxSvAction)

        RecDXFSvAction = QAction('Сохранить приемники в dxf', self)
        #SaveInDXFAction.setShortcut('Ctrl+R')
        RecDXFSvAction.triggered.connect(self.SaveRecInDXF) 
        ResultMenu.addAction(RecDXFSvAction)

        

        settingsMenu = menubar.addMenu('&Настройки')

        self.fl_state = "fl32"

        FloatTypeGroup = QActionGroup(self)
        FloatTypeGroup.setExclusive(True)
        fl32 = QAction('float32', self)
        fl32.setCheckable(True)
        fl32.setChecked(True)
        fl64 = QAction('float64', self)
        fl64.setCheckable(True)
        FloatTypeGroup.addAction(fl32)
        FloatTypeGroup.addAction(fl64)
        FloatTypeGroup.triggered.connect(lambda x:setTypes(x.text()))

        sep1 = QAction('', self) 
        sep1.setSeparator(True)

        dragable = QAction('Передвигаемость', self) 
        dragable.setCheckable(True)
        self.drag_state = True
        dragable.setChecked(True)
        dragable.triggered.connect(self.DragObj)

        
        settingsMenu.addAction(fl32)
        settingsMenu.addAction(fl64)
        settingsMenu.addAction(sep1)
        settingsMenu.addAction(dragable)

        Splitter1 = QSplitter(Qt.Horizontal) 
        Splitter1.addWidget(LeftPanelFrame)
        Splitter1.addWidget(SceneFrame)
        Splitter1.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        Splitter1.setStretchFactor(1, 4) 
   
        vbox = QVBoxLayout(self)
        vbox.addWidget(Splitter1)
        self.central_widget = QWidget()
        self.central_widget.setLayout(vbox)
        self.setCentralWidget(self.central_widget) 

        # Dict for using cashe of calculate
        try:
            with open('Cashe_files/cahse.pkl', "rb" ) as f:
                self.Cashe, self.MutualCashe = pickle.load(f)
        except Exception:
            self.Cashe = {}
            self.MutualCashe = {}

    def DragObj(self,state):
        self.drag_state = state
        self.scene.dragable(state)

    def closeEvent(self, event):
        Message = QMessageBox(QMessageBox.Question,  'Выход из программы',
            "Вы дейстивлеьно хотите выйти?", parent=self)
        Message.addButton('Да', QMessageBox.YesRole)
        Message.addButton('Нет', QMessageBox.NoRole)
        #Message.addButton('Сохранить', QMessageBox.ActionRole)
        reply = Message.exec()
        if reply == 0:
            with open('Cashe_files/cahse.pkl', "wb" ) as f:
                pickle.dump([self.Cashe,self.MutualCashe],f)
            event.accept()
        elif reply == 1:
            event.ignore()

    def on_mouse_move(self,click):
        x, y = click.xdata, click.ydata
        if self.last_fig == 'Mfield':
            if self.rbf is not None and x is not None and y is not None:
                x, y = np.around(x, 3), np.around(y, 3)
                H  = np.around(self.rbf(x,y))
                self.status_string.showMessage(f"x: {x}, y: {y}, H: {H}")
        elif self.last_fig == 'Irec':
            if x is not None and y is not None:
                x, y = np.around(x, 3), np.around(y, 3)
                self.status_string.showMessage(f"x: {x}, y: {y}")

    def SavePlot(self):
        try:
            fname = QFileDialog.getSaveFileName(self, 'Сохранить файл', self.path_home,'*.png;;*.png;;*.pdf;;*.svg;;*.eps')

            #self.fig.set_size_inches(4, 4,forward=True) # Изменяем размер сохраняемого графика
            self.fig.savefig(fname[0], format=fname[1][2:], dpi=300) # Cохраняем графики
        except Exception as ex:
            print(ex)
      
    def SetColorTContour(self,row,col):
        if col == 2:
            color = QColorDialog.getColor(initial=self.table_contour.item(row,col).background().color(),\
                                            parent=self,title="Цвет контура",options=QColorDialog.ShowAlphaChannel)

            if color.isValid():
                self.table_contour.item(row,2).setBackground(color)
                self.SetCountur(row,1)

    def SetCountur(self,row,col,tc=None):
        state = self.table_contour.item(row,0).checkState()
        sp_levels = []
        #ax1.clabel(cs)
        #countur_date = cs.allsegs

        if state == Qt.Checked:
            if col==0:
                self.d_contur[row][0] = True

            try:
                if col==0 or col==1:
                    c = self.table_contour.item(row,2).background().color()
                    l = self.table_contour.item(row,1).text()

                    if self.d_contur[row][1] is None and l!='' and self.make_contur is not None:
                        self.d_contur[row][1] = self.make_contur(float(l),c.name())  #= lambda l, c: ax1.contour(xi, zi, hi, levels = l, colors = c )

                        if tc is None:
                            self.Canv.draw()
                        elif tc == 1:
                            sp_levels.append([l,(c.red(),c.green(),c.blue()),self.d_contur[row][1].allsegs])


                    elif self.d_contur[row][1] is not None and l!='' and self.make_contur is not None:
                        ds = self.make_contur(float(l),c.name())

                        if tc is None:
                            self.Canv.draw()

                        for c in self.d_contur[row][1].collections:
                            c.remove()
                        
                        if tc is None:
                            self.Canv.draw()
                        elif tc == 1:
                            sp_levels.append([l,(c.red(),c.green(),c.blue()),ds.allsegs])

                        self.d_contur[row][1] = ds

            except Exception as ex:
                print(ex)
                 
        elif state == Qt.Unchecked and col==0:
            self.d_contur[row][0] = False
            if self.d_contur[row][1] is not None:
                for c in self.d_contur[row][1].collections:
                    c.remove()
                self.d_contur[row][1] = None
                self.Canv.draw()

        if tc == 1:
            return sp_levels 


    def ResizeTablesContour(self):
        try:
            new_rows = self.tbc_rows.value()
            
            if new_rows>self.old_rows:
                #x = float(self.table_cord.item(self.old_rows-1,0).text())
                #y = float(self.table_cord.item(self.old_rows-1,1).text())
                self.table_contour.setRowCount(new_rows)
                for i in range(self.old_rows,new_rows):
                    self.d_contur[i] = [False, None]
                    item = QTableWidgetItem("")
                    item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
                    item.setCheckState(Qt.Unchecked)
                    self.table_contour.setItem(i,0, item)
                    self.table_contour.setItem(i,1, QTableWidgetItem(""))
                    self.table_contour.setItem(i,2, QTableWidgetItem(""))
                    self.table_contour.item(i,2).setBackground(QColor(255,0,0))
                     
            elif new_rows<self.old_rows:
                for i in range(new_rows,self.old_rows):
                    self.table_contour.item(i,0).setCheckState(Qt.Unchecked)
                    del self.d_contur[i]

                self.table_contour.setRowCount(new_rows)


            self.old_rows = new_rows
        except Exception as ex:
            print(ex)

    def ShowCanvas(self,SpaceCord,H_area,area, cashe=None, sp_levels=None, gen=None):
        try:
            if cashe is not None:
                numb = len(self.Cashe)+1
                self.Cashe[cashe] = numb
                np.save(f'Cashe_files/XYZ_{numb}.npy',SpaceCord)
                np.save(f'Cashe_files/H_{numb}.npy',H_area)

            if sp_levels is None:
                self.Ind.setLabelText("Создаётся график")
                self.Ind.reset()
            else:
                self.Ind.setLabelText("Создаются линии уровня")

            self.last_fig = 'Mfield'

            tp, area_calc, z = area[:3]
            print(np.shape(H_area))
            H_area = np.linalg.norm(H_area,axis=1)
            self.fig.clear()
            #sp_levels = []
        
            if tp == "horizontal_area":
                X,Y = SpaceCord[0,:],SpaceCord[1,:]

                xi, yi = np.linspace(X.min(), X.max(), 100), np.linspace(Y.min(), Y.max(), 100)
                xi, yi = np.meshgrid(xi, yi)

                self.rbf = scipy.interpolate.Rbf(X, Y, H_area, function='linear') #'inverse'
                zi = self.rbf(xi, yi)

                ax = self.fig.add_subplot(111) #
                

                im=ax.imshow(zi, vmin=H_area.min(), vmax=H_area.max(), cmap='jet', norm=LogNorm(), origin='lower',extent=[X.min(), X.max(), Y.min(), Y.max()]) #H_area.max()
                cl = self.fig.colorbar(im)
                cl.set_label('H, А/м',verticalalignment = "top", x=-20) #, rotation=270 #,position=(20,0)

                self.make_contur = lambda l, c: ax.contour(xi, yi, zi, levels = l, colors = c )
                sp = []
                for k in self.d_contur:
                    self.d_contur[k][1]=None
                    if sp_levels is None:
                        self.SetCountur(k,0)
                    else:
                        sp+=self.SetCountur(k,0,tc=1)
                        
                        #print(len(sp_levels))
                if sp_levels is not None:       
                    sp_levels.append(sp)
                
                
                ax.set_xlabel(u'X, м') # Подпись оси х ,fontsize=self.shr_gr
                ax.set_ylabel(u'Y, м') # Подпись оси у ,fontsize=self.shr_gr
                #ax.set_title("Тест\n") #,fontsize=self.shr_gr

                if sp_levels is None:
                    self.Canv.draw() # Выводим график в виджет
                    self.tab.setCurrentIndex(1) # Делаем активной созданую закладку
                elif sp_levels is not None and gen is not None:
                    gen()

                """ elif tc == 1:
                    return sp_levels """

            
            elif tp == "vertical_area" and sp_levels is None:
                X,Y,Z = SpaceCord[0,:],SpaceCord[1,:],SpaceCord[2,:]

                if area_calc[2]-area_calc[0]>=area_calc[3]-area_calc[1]:
                    
                    ax1 = self.fig.gca()
                    
                    
                    # Создание графика
                    xi, zi = np.linspace(X.min(), X.max(), 100), np.linspace(Z.min(), Z.max(), 100)
                    xi, zi = np.meshgrid(xi, zi)

                    self.rbf = scipy.interpolate.Rbf(X, Z, H_area, function='linear')
                    hi = self.rbf(xi, zi)

                    im = ax1.imshow(hi, vmin=H_area.min(), vmax=H_area.max(), cmap='jet', norm=LogNorm(), origin='lower',extent=[X.min(), X.max(), Z.min(), Z.max()],aspect="auto") #
                    self.make_contur = lambda l, c: ax1.contour(xi, zi, hi, levels = l, colors = c )

                    for k in self.d_contur:
                        self.d_contur[k][1]=None
                        self.SetCountur(k,0)

                    if area_calc[3]-area_calc[1]>0.05:
                        ax2 = ax1.twiny()
                        ax2.set_xlim([Y.min(),Y.max()])
                        ax2.set_xlabel(u'Y, м\n') # Подпись оси у ,fontsize=self.shr_gr
                
                    ax1.set_xlabel(u'X, м') # Подпись оси х ,fontsize=self.shr_gr
                    ax1.set_ylabel(u'Z, м') # Подпись оси у ,fontsize=self.shr_gr
                    
                else:
                    if area_calc[2]-area_calc[0]>0.05:
                        ax1 = self.fig.gca()
                        ax2 = ax1.twiny()
                    else:
                        ax2 = self.fig.gca()

                    xi, zi = np.linspace(Y.min(), Y.max(), 100), np.linspace(Z.min(), Z.max(), 100)
                    xi, zi = np.meshgrid(xi, zi)

                    self.rbf = scipy.interpolate.Rbf(Y, Z, H_area, function='linear')
                    hi = self.rbf(xi, zi)

                    im = ax2.imshow(hi, vmin=H_area.min(), vmax=H_area.max(), origin='lower',
                            extent=[Y.min(), Y.max(), Z.min(), Z.max()],aspect="auto")
                    self.make_contur = lambda l, c: ax2.contour(xi, zi, hi, levels = l, colors = c)

                    for k in self.d_contur:
                        self.d_contur[k][1]=None
                        self.SetCountur(k,0)
                    
                    if area_calc[2]-area_calc[0]>0.05:
                        ax1.set_xlim([X.min(),X.max()])
                        ax1.set_xlabel(u'X, м') 
                    
                    ax2.set_xlabel(u'Y, м\n')
                    ax2.set_ylabel(u'Z, м') # Подпись оси у ,fontsize=self.shr_gr

                cl = self.fig.colorbar(im)
                cl.set_label('H, А/м',verticalalignment = "top", x=-20) #, rotation=270 #,position=(20,0)

                self.Canv.draw() # Выводим график в виджет
                self.tab.setCurrentIndex(1)

            #self.Ind.close()
        except Exception as ex:
            print("ShowCanvas",ex)

    def ProgresCalc(self):
        Ind = QProgressDialog('Производится расчёт','Отмена', 0, 0, self)
        Ind.setWindowTitle('Расчет ')
        Ind.setMinimumDuration(0)
        Ind.setWindowModality(Qt.WindowModal)
        Ind.setAutoClose(True)
        Ind.show()
        return Ind

    def createSignals(self):
        setNewSignal = lambda x: self.Ind.setValue(x)
        setStopSignal = lambda x: self.Ind.canceled.connect(x)
        setRange = lambda x,y: self.Ind.setRange(x,y)
        return (setStopSignal,setRange,setNewSignal)     

    def get_object_data(self):
        try:
            lst = []
            layers = {}
            for item, objs in self.SoursesObjectsChildren.items():
                t = item.text(1)
                layers[t] = set()
                for key in objs:
                    obj = self.SoursesObjectsDict[key]
                    if obj.type_link=="object" and obj.state:
                        data = obj.graphic_item.menu.data.read_data()
                        lst.append(data)
                        layers[t].add(data)
                            

            #print(layers)
            return tuple(sorted(lst)), layers
        except Exception as ex:
            print(ex)

    def Run_area(self):
        tree= self.AreasObjectsTree
        ObjDict, ObjChildren = self.AreasObjectsDict, self.AreasObjectsChildren
        item = tree.selectedItems()
        if len(item) != 1: return
        else: item = item[0]

        print(item)

        obj = ObjDict[item]
        if obj.type_link != "object" or (obj.type_link == "object" and obj.type_object == "one_point"):
            return

        area = obj.graphic_item.menu.data.read_data()
        sourses,layers = self.get_object_data()

        self.Ind = self.ProgresCalc()
        print("start")
        dl = self.dl.value()
        da = self.da.value()

        cashe = (sourses,area,dl,da)

        if cashe not in self.Cashe:
            run_area_calc(sourses,area,da,dl, callback_func = (lambda S, H :self.ShowCanvas(S,H,area,cashe=cashe),self.createSignals(),None))#self.Ind
        else:
            numb = self.Cashe[cashe]
            S = np.load(f'Cashe_files/XYZ_{numb}.npy')
            H = np.load(f'Cashe_files/H_{numb}.npy')
            self.ShowCanvas(S,H,area)
            print("end_cashe")


    def PointCalc(self, who, data=None):
        try:
            if who == "sourses":
                sourses,layers = self.get_object_data()

                points = []
                for obj in self.AreasObjectsDict.values():
                    if obj.type_link=="object":
                        if obj.type_object == "one_point" and obj.state:
                            points.append(obj.graphic_item.menu.data)

                point_calc(sourses,points, self.da.value(), self.dl.value())
            
            elif who == "areas":
                sourses,layers = self.get_object_data()
                points = [data]

                point_calc(sourses,points, self.da.value(), self.dl.value())

        except Exception:
            pass

        #print("end_move")


    def RecCalc(self):
        try:
            
            receivers = []
            for key, val in self.ReceiverObjectsDict.items():
                if key.checkState(0) == Qt.Checked and val.type_link == "object":
                    receivers.append(val.graphic_item.menu.data.read_data())

            sourses = self.get_object_data()[0]

            dl = self.dl.value()
            da = self.da.value()
            
            Id, Lines, setTextPos, get_colors,self.ExcSv = receivers_calc(sourses, receivers, dl, da, self.MutualCashe, excel=True)
            self.RecCanvas(Id, Lines, setTextPos, get_colors)

            print("RecCalc")
        except Exception as ex:
            print(ex, "RecCalc")

    def RecCanvas(self, Id, Lines, setTextPos, get_colors):
        self.last_fig = 'Irec'
        fig_lines = []
        minXL, maxXL, minYL, maxYL = float("inf"), float("-inf"), float("inf"), float("-inf")
        for i in Lines:
            minXL, maxXL, minYL, maxYL = min(i[0][0],i[1][0],minXL), max(i[0][0],i[1][0],maxXL),min(i[0][1],i[1][1],minYL), max(i[0][1],i[1][1],maxYL)
            fig_lines.append([i[0][:-1],i[1][:-1]])

        self.fig.clear()
        ax = self.fig.add_subplot(111) #

        border = abs(((maxXL-minXL)+(maxYL-minYL))/2*0.1)

        ax.set_xlim(minXL-border, maxXL+border)
        ax.set_ylim(minYL-border, maxYL+border)

        line_segments = LineCollection(fig_lines,
                                    linewidths=2,
                                    linestyles='solid',cmap=jet) #, alpha=0.4 linewidths=(0.5, 1, 1.5, 2)

        line_segments.set_array(Id)
        
        for ((x,y), deg, ha),i in zip(setTextPos(fig_lines),Id):
            ax.text(x,y,str(round(i,2)),\
                        horizontalalignment=ha,verticalalignment="bottom",rotation=deg,fontsize=8)#fontproperties=arial_font,fontsize=text

        lim = (np.min(Id), np.max(Id))
        line_segments.set_clim(lim)

        #print(get_colors(Id,lim,jet,Normalize))

        ax.add_collection(line_segments)

        axcb = self.fig.colorbar(line_segments)
        axcb.set_label('I, A')

        ax.set_title('Ток в приемниках')
        ax.set_xlabel(u'X, м') 
        ax.set_ylabel(u'Y, м') 
        ax.grid()

        self.Canv.draw() # Выводим график в виджет
        self.tab.setCurrentIndex(1)

    def SaveExcel(self):
        try:
            wb = Workbook()
            ws=wb.active
            fname = QFileDialog.getSaveFileName(self, 'Сохранить файл', self.path_home,'*.xlsx;;*.xls')[0]
            for (i,j),val in self.ExcSv.items():
                ws.cell(row=i+1,column=j+1).value = val
            wb.save(fname)
        except Exception as ex:
            print(ex)

    def SaveRecInDXF(self):
        try:
            fname = QFileDialog.getSaveFileName(self, 'Сохранить файл', self.path_home,'*.dxf')[0]
            receivers = []
            for key, val in self.ReceiverObjectsDict.items():
                if key.checkState(0) == Qt.Checked and val.type_link == "object":
                    receivers.append(val.graphic_item.menu.data.read_data())

            sourses = self.get_object_data()[0]

            dl = self.dl.value()
            da = self.da.value()
            
            Id, Lines, setTextPos, get_colors = receivers_calc(sourses, receivers, dl, da, self.MutualCashe)

            lim = (np.min(Id), np.max(Id))
            colors = get_colors(Id,lim,jet,Normalize)
            text = [str(round(i,2)) for i in Id]

            dxf.SaveRecInDXF(Lines,text,colors,setTextPos,fname)

        except Exception as ex:
            print("SaveRecInDXF",ex)

    def SaveInDXFstart(self):
        try:
            self.w = Save_Widget(self.AreasObjectsDict,self.AreasObjectsChildren,self.SaveInDXFfinish)
            self.w.show()
        except Exception as ex:
            print(ex)

    
    def SaveInDXFfinish(self,areas,area_layers,tsf):
        try:
            fname = QFileDialog.getSaveFileName(self, 'Сохранить файл', self.path_home,'*.dxf')[0]
            sourses,sourses_layers = self.get_object_data()
            self.Ind = self.ProgresCalc()

            dl = self.dl.value()
            da = self.da.value()

            h_areas = [i for i in areas if i[0] == "horizontal_area"]
            lvr_levels = [area_layers[i] for i in areas if i[0] == "horizontal_area"]


            p_areas = [i for i in areas if i[0] == "one_point"]
            lvr_points = [area_layers[i] for i in areas if i[0] == "one_point"]

            sp_points = point_calc(sourses, p_areas, da, dl, dxf=True)
            

            self.sp_levels = []
            f_l = []
            for area in h_areas:
                cashe = (sourses,area,dl,da)
                if area[0] == "horizontal_area":
                    if cashe not in self.Cashe:
                        f = (lambda s,a,da,dl,ch,lv,sg,ordr: (lambda :run_area_calc(s,a,da,dl,\
                            callback_func = (lambda S, H :self.ShowCanvas(S,H,a,\
                            cashe=ch,sp_levels=lv),sg,ordr))))(sourses,area,da,dl,cashe,self.sp_levels,self.createSignals(),self.CalcOrder)
                        f_l.append(f)
                    else:
                        numb = self.Cashe[cashe]
                        S = np.load(f'Cashe_files/XYZ_{numb}.npy')
                        H = np.load(f'Cashe_files/H_{numb}.npy')
                        f = (lambda s,h,a,lv,ordr: lambda: (self.ShowCanvas(s,h,a,sp_levels=lv,gen=ordr)))(S,H,area,self.sp_levels,self.CalcOrder)
                        f_l.append(f)
                    print("yes")
            def end(a,b,c,d,e):
                dxf.SaveInDXF(a,b,c,d)
                e.reset()
            #f = lambda: dxf.SaveInDXF(sp_points,self.sp_levels,sourses,(fname,tsf))
            f = lambda: end((lvr_points,sp_points),(lvr_levels,self.sp_levels),sourses_layers,(fname,tsf),self.Ind)
            f_l.append(f)
            self.gen = (i for i in f_l)
            self.CalcOrder()
        except Exception as ex:
            print('aaaaa',ex)

    def CalcOrder(self):
        next(self.gen)()
       
            
    def ChangeCheckBox(self,item,colum,type_tree):
        """ Check state control """
        try:
            if type_tree == "sourses":
                if colum == 0:
                    t = self.SoursesObjectsDict[item].type_link
                    state = item.checkState(0)
                    if t == "object":
                        if state == Qt.Checked:
                            self.scene.addItem(self.SoursesObjectsDict[item].graphic_item)
                            self.SoursesObjectsDict[item].state= True
                        else:
                            self.scene.removeItem(self.SoursesObjectsDict[item].graphic_item)
                            self.SoursesObjectsDict[item].state = False
                    elif t == "layer":
                        if state == Qt.Checked:
                            for child in self.SoursesObjectsChildren[item]:
                                child.setCheckState(0, Qt.Checked)
                        else:
                            for child in self.SoursesObjectsChildren[item]:
                                child.setCheckState(0, Qt.Unchecked)

            elif type_tree == "areas":
                if colum == 0:
                    t = self.AreasObjectsDict[item].type_link
                    state = item.checkState(0)
                    if t == "object":
                        if state == Qt.Checked:
                            self.scene.addItem(self.AreasObjectsDict[item].graphic_item)
                            self.AreasObjectsDict[item].state = True
                        else:
                            self.scene.removeItem(self.AreasObjectsDict[item].graphic_item)
                            self.AreasObjectsDict[item].state = False
                    elif t == "layer":
                        if state == Qt.Checked:
                            for child in self.AreasObjectsChildren[item]:
                                child.setCheckState(0, Qt.Checked)
                        else:
                            for child in self.AreasObjectsChildren[item]:
                                child.setCheckState(0, Qt.Unchecked)

            elif type_tree == "receivers":
                if colum == 0:
                    t = self.ReceiverObjectsDict[item].type_link
                    state = item.checkState(0)
                    if t == "object":
                        if state == Qt.Checked:
                            self.scene.addItem(self.ReceiverObjectsDict[item].graphic_item)
                            self.ReceiverObjectsDict[item].state = True
                        else:
                            self.scene.removeItem(self.ReceiverObjectsDict[item].graphic_item)
                            self.ReceiverObjectsDict[item].state = False
                    elif t == "layer":
                        if state == Qt.Checked:
                            for child in self.ReceiverObjectsChildren[item]:
                                child.setCheckState(0, Qt.Checked)
                        else:
                            for child in self.ReceiverObjectsChildren[item]:
                                child.setCheckState(0, Qt.Unchecked)
        except Exception as ex:
            print(ex,"ChangeCheckBox")
                    
    
    def OpenObjMenu(self,item,colum,type_tree):
        if type_tree == "sourses": obj = self.SoursesObjectsDict
        elif type_tree == "areas": obj = self.AreasObjectsDict
        elif type_tree == "receivers": obj = self.ReceiverObjectsDict

        if obj[item].type_link == "object" and colum == 1:
            obj[item].graphic_item.menu.show()

    def AddObj(self, type_tree, type_obj):
        try:
            self.BlockSignals(True)

            if type_tree == "sourses": 
                tree, link = self.SoursesObjectsTree, "images/sourse.png"
                ObjDict, ObjChildren = self.SoursesObjectsDict, self.SoursesObjectsChildren
            elif type_tree == "areas": 
                tree, link = self.AreasObjectsTree, "images/area.png"
                ObjDict, ObjChildren = self.AreasObjectsDict, self.AreasObjectsChildren
            elif type_tree == "receivers": 
                tree, link = self.ReceiverObjectsTree, "images/receiver.png"
                ObjDict, ObjChildren = self.ReceiverObjectsDict, self.ReceiverObjectsChildren


            item = tree.selectedItems()
            if type_obj =="layer" : item = None
            elif len(item)>1: return
            else: item = item[0]

            if type_obj =="layer":
                parent = TreeWidgetItem(tree, ["","new_layer"])
                parent.setFlags(parent.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsEditable)
                parent.setCheckState(0, Qt.Unchecked)
                parent.setIcon(1,QIcon(link))

                ObjDict[parent] = Atributs()
                ObjChildren[parent] = set()

            else:
                rect  = self.view.mapToScene(self.view.rect()).boundingRect()
                wd, hg = rect.width()/4, rect.height()/4
                cx,cy = rect.center().x(),rect.center().y()

                if ObjDict[item].type_link != "layer": item = item.parent()

                if type_obj == "reactor": 
                    DrawObj = GraphicsCircleItem((cx,cy,wd),calc_func=self.PointCalc if self.pcalc else None)
                    link_obj = "images/reactor.png"
                elif type_obj == "conductor": 
                    DrawObj = GraphicsPolylineItem([[cx-wd,cy-hg],[cx+wd,cy+hg]],calc_func=self.PointCalc if self.pcalc else None)
                    link_obj = "images/conductor.png"
                elif type_obj == "horizontal_area": 
                    DrawObj = GraphicsRectItem((cx-wd,cy-hg,cx+wd,cy+hg))
                    link_obj = "images/rectangle.png"
                elif type_obj == "vertical_area": 
                    DrawObj = GraphicsLineItem((cx-wd,cy-hg,cx+wd,cy+hg))
                    link_obj = "images/line.png" 
                elif type_obj == "one_point": 
                    DrawObj = OneCalcCircle((cx,cy), calc_func=self.PointCalc if self.pcalc else None)
                    link_obj = "images/point.png"
                elif type_obj == "recconductor": 
                    DrawObj = RecPolylineItem([[cx-wd,cy-hg],[cx+wd,cy+hg]])
                    link_obj = "images/conductor.png"

                child = TreeWidgetItem(["","new_obj"])
                child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
                child.setCheckState(0, Qt.Unchecked)
                child.setIcon(1,QIcon(link_obj))
                item.addChild(child)

                DrawObj.hndl*=1/self.view.current_scale  
                DrawObj.updateHandlesPos()
                DrawObj.dragable(self.drag_state)
                DrawObj.menu.data.name = "new_obj"
                DrawObj.menu.setListName = self.Callbackname(child,DrawObj)

                ObjDict[child] = Atributs(type_link="object",type_object=type_obj, graphic_item=DrawObj)  
                ObjChildren[item].add(child)

            self.BlockSignals(False)
        except Exception as ex:
            print("AddObj",ex)

    
    def DelObj(self,type_tree):
        try:
            if type_tree == "sourses": 
                tree= self.SoursesObjectsTree
                ObjDict, ObjChildren = self.SoursesObjectsDict, self.SoursesObjectsChildren
                item = tree.selectedItems()
                if len(item)>1: return
                else: item = item[0]

                if ObjDict[item].type_link == "layer":
                    message_text = f'Вы действительно желаете удалить из "источников" cлой "{item.text(1)}" со всеми его объектами?'
                elif ObjDict[item].type_link == "object":
                    message_text = f'Вы действительно желаете удалить из "источников" объект "{item.text(1)}"?'

            elif type_tree == "areas": 
                tree= self.AreasObjectsTree
                ObjDict, ObjChildren = self.AreasObjectsDict, self.AreasObjectsChildren
                item = tree.selectedItems()
                if len(item)>1: return
                else: item = item[0]

                if ObjDict[item].type_link == "layer":
                    message_text = f'Вы действительно желаете удалить из "областей" cлой "{item.text(1)}" со всеми его объектами?'
                elif ObjDict[item].type_link == "object":
                    message_text = f'Вы действительно желаете удалить из "областей" объект "{item.text(1)}"?'

            elif type_tree == "receivers": 
                tree = self.ReceiverObjectsTree
                ObjDict, ObjChildren = self.ReceiverObjectsDict, self.ReceiverObjectsChildren
                item = tree.selectedItems()
                if len(item)>1: return
                else: item = item[0]

                if ObjDict[item].type_link == "layer":
                    message_text = f'Вы действительно желаете удалить из "приёмников" cлой "{item.text(1)}" со всеми его объектами?'
                elif ObjDict[item].type_link == "object":
                    message_text = f'Вы действительно желаете удалить из "приёмников" объект "{item.text(1)}"?'

            Message = QMessageBox(QMessageBox.Question,  'Удаление',
                message_text, parent = self)
            Message.addButton('Да', QMessageBox.YesRole)
            Message.addButton('Нет', QMessageBox.NoRole)
            reply = Message.exec()       
            if reply == 0:
                if ObjDict[item].type_link == "layer":
                    for it in ObjChildren[item]:
                        if ObjDict[it].state:
                            self.scene.removeItem(ObjDict[it].graphic_item)
                        del ObjDict[it]

                    tree.takeTopLevelItem(tree.indexOfTopLevelItem(item))
                    del ObjDict[item]
                    del ObjChildren[item]
                
                elif ObjDict[item].type_link == "object":
                    if ObjDict[item].state:
                        self.scene.removeItem(ObjDict[item].graphic_item)
                    
                    del ObjDict[item]
                    ObjChildren[item.parent()].remove(item)
                    parent = item.parent()
                    parent.takeChild(parent.indexOfChild(item))

        except Exception as ex:
            print("DelObj", ex)


    def BlockSignals(self,trig):
        self.SoursesObjectsTree.blockSignals(trig)
        self.AreasObjectsTree.blockSignals(trig)
        self.ReceiverObjectsTree.blockSignals(trig)

    def ClearTrees(self):
        self.SoursesObjectsTree.clear()
        self.AreasObjectsTree.clear()
        self.ReceiverObjectsTree.clear()

        self.SoursesObjectsDict.clear()
        self.AreasObjectsDict.clear()
        self.ReceiverObjectsDict.clear()

        self.SoursesObjectsChildren.clear()
        self.AreasObjectsChildren.clear()
        self.ReceiverObjectsChildren.clear()

        self.scene.clear()

    @staticmethod
    def Callbackname(item, draw):
        return lambda: item.setText(1,draw.menu.data.name)

    def LoadCalcData(self):
        fname = QFileDialog.getOpenFileName(self, 'Открыть файл', self.path_home,'*.mfc')
        if fname[0] == "" and  fname[1] == "": return
        fname = fname[0]

        with open(fname, "r", encoding="utf8") as f:
            data  = json.load(f)

        name = os.path.splitext(os.path.split(fname)[1])[0]

        self.ClearTrees()

        # Block signals for all metods
        self.BlockSignals(True)

        self.view.scale(1/self.view.current_scale,1/self.view.current_scale)

        wV = self.view.size().width()
        hV = self.view.size().height()

        sourses, areas, receivers, settings = data["Sourses"], data["Areas"], data.get("Receivers",[]), data["Settings"]
        xmin, ymin, xmax, ymax  = float("inf"), float("inf"), -float("inf"), -float("inf") 
        fc = lambda a,b: (min(a[0],b[0]),min(a[1],b[1]),max(a[2],b[2]),max(a[3],b[3]))

        self.dl.setValue(settings.get("dl",0.01))
        self.da.setValue(settings.get("da",1))

        Sourses = []
        Areas = []
        Receivers = []

        for name_lv, obj in sourses:
            l = [name_lv, []]
            for data in obj:
                data_obj = MenuData(init_data=data)
                if data_obj.type_object == "reactor":
                    cord = ((data_obj.X.number_gui, -data_obj.Y.number_gui), data_obj.Rnar.number_gui)
                    l[1].append([data_obj.name,data_obj.type_object,cord,data_obj])
                    xmin, ymin, xmax, ymax = fc(data_obj.borders(),(xmin, ymin, xmax, ymax))

                elif data_obj.type_object == "conductor":
                    cord = [[x, -y] for x, y in data_obj.tbl_XY.number_gui]
                    l[1].append([data_obj.name,data_obj.type_object,cord,data_obj])
                    xmin, ymin, xmax, ymax = fc(data_obj.borders(),(xmin, ymin, xmax, ymax))

            Sourses.append(l)

        for name_lv, obj in areas:
            l = [name_lv, []]
            for data in obj:
                data_obj = MenuData(init_data=data)
                if data_obj.type_object == "horizontal_area":
                    cord = (data_obj.X1.number_gui, -data_obj.Y2.number_gui, data_obj.X2.number_gui, -data_obj.Y1.number_gui)
                    l[1].append([data_obj.name,data_obj.type_object,cord,data_obj])
                    xmin, ymin, xmax, ymax = fc(data_obj.borders(),(xmin, ymin, xmax, ymax))

                elif data_obj.type_object == "vertical_area":
                    cord = (data_obj.X1.number_gui, -data_obj.Y1.number_gui, data_obj.X2.number_gui, -data_obj.Y2.number_gui)
                    l[1].append([data_obj.name,data_obj.type_object,cord,data_obj])
                    xmin, ymin, xmax, ymax = fc(data_obj.borders(),(xmin, ymin, xmax, ymax))

                elif data_obj.type_object == "one_point":
                    cord = (data_obj.X.number_gui, -data_obj.Y.number_gui)
                    l[1].append([data_obj.name,data_obj.type_object,cord,data_obj])
                    xmin, ymin, xmax, ymax = fc(data_obj.borders(),(xmin, ymin, xmax, ymax))

            Areas.append(l)


        for name_lv, obj in receivers:
            l = [name_lv, []]
            for data in obj:
                data_obj = MenuData(init_data=data)
                if data_obj.type_object == "recconductor":
                    cord = [[x, -y] for x, y in data_obj.tbl_XY.number_gui]
                    l[1].append([data_obj.name,data_obj.type_object,cord,data_obj])
                    xmin, ymin, xmax, ymax = fc(data_obj.borders(),(xmin, ymin, xmax, ymax))

            Receivers.append(l)


        scl_layers = min(wV/(xmax-xmin),hV/(ymax-ymin)) if xmax>-float("inf") and xmin<float("inf") and ymax>-float("inf") and ymin<float("inf") else 1
        self.LoadData(Sourses, Areas, Receivers, scl_layers)            
           
        self.view.current_scale = scl_layers
        self.view.scale(scl_layers,scl_layers)
        if xmax>-float("inf") and xmin<float("inf") and ymax>-float("inf") and ymin<float("inf"):
            self.view.centerOn(0.5*(xmin+xmax),-0.5*(ymin+ymax))

        # Unblock signals for all metods
        self.BlockSignals(False)

    def load_dxf_file(self):
        # Create file dialog for open dxf file and get its name and path
        fname = QFileDialog.getOpenFileName(self, 'Открыть файл', self.path_home,'*.dxf') # Обрати внимание на последний элемент
        if fname[0] == "" and  fname[1] == "": return
        fname = fname[0]
        
        name = os.path.splitext(os.path.split(fname)[1])[0]

        self.ClearTrees()

        # Block signals for all metods
        self.BlockSignals(True)

        self.view.scale(1/self.view.current_scale,1/self.view.current_scale)

        layers = dxf.OpenFile(fname)


        wV = self.view.size().width()
        hV = self.view.size().height()

        xmin, ymin, xmax, ymax  = float("inf"), float("inf"), -float("inf"), -float("inf") 
        fc = lambda a,b: (min(a[0],b[0]),min(a[1],b[1]),max(a[2],b[2]),max(a[3],b[3]))
         
        Sourses = []
        Areas = []
        Receivers = []

        for layer in layers:
            
            x1,y1,x2,y2 = layers[layer]["size"]
            xmin, ymin, xmax, ymax = fc((x1,y1,x2,y2),(xmin, ymin, xmax, ymax))

            if layers[layer]["type"] == 'objects':
                l = [layers[layer]["name"],[]]
                for name_obj in layers[layer]["circle"]:
                    l[1].append([name_obj,"reactor",layers[layer]["circle"][name_obj],None])
                for name_obj in layers[layer]["lwpolyline"]:
                    l[1].append([name_obj,"conductor",layers[layer]["lwpolyline"][name_obj],None])
                Sourses.append(l)

            elif layers[layer]["type"] == 'areas':
                l = [layers[layer]["name"],[]]
                for name_obj in layers[layer]["rectangle"]:
                    l[1].append([name_obj,"horizontal_area",layers[layer]["rectangle"][name_obj],None])
                for name_obj in layers[layer]["line"]:
                    l[1].append([name_obj,"vertical_area",layers[layer]["line"][name_obj],None])
                for name_obj in layers[layer]["point"]:
                    l[1].append([name_obj,"one_point",layers[layer]["point"][name_obj],None])
                Areas.append(l)

            elif layers[layer]["type"] == 'receivers':
                l = [layers[layer]["name"],[]]
                for name_obj in layers[layer]["lwpolyline"]:
                    l[1].append([name_obj,"recconductor",layers[layer]["lwpolyline"][name_obj],None])
                Receivers.append(l)


        scl_layers = min(wV/(xmax-xmin),hV/(ymax-ymin)) if xmax>-float("inf") and xmin<float("inf") and ymax>-float("inf") and ymin<float("inf") else 1

        self.LoadData(Sourses, Areas, Receivers, scl_layers)            

        self.view.current_scale = scl_layers  
        self.view.scale(scl_layers,scl_layers)
        if xmax>-float("inf") and xmin<float("inf") and ymax>-float("inf") and ymin<float("inf"):
            self.view.centerOn(0.5*(xmin+xmax),-0.5*(ymin+ymax))

        # Unblock signals for all metods
        self.BlockSignals(False)

    def LoadData(self, Sourses, Areas, Receivers, scl_layers):
        for name_lv, obj in Sourses:
            parent = TreeWidgetItem(self.SoursesObjectsTree, ["",name_lv])
            parent.setFlags(parent.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsEditable)
            parent.setCheckState(0, Qt.Unchecked)
            parent.setIcon(1,QIcon("images/sourse.png"))

            self.SoursesObjectsDict[parent] = Atributs()
            self.SoursesObjectsChildren[parent] = set()

            for name_obj, tp, cord, data in obj:
                if tp == "reactor":
                    child = TreeWidgetItem(["",name_obj])
                    child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
                    child.setCheckState(0, Qt.Unchecked)
                    child.setIcon(1,QIcon("images/reactor.png"))
                    parent.addChild(child)

                    Circle = GraphicsCircleItem((cord[0][0],-cord[0][1],cord[1]),data=data,calc_func=self.PointCalc if self.pcalc else None)
                    Circle.hndl*=1/scl_layers  
                    Circle.updateHandlesPos()
                    Circle.dragable(self.drag_state)
                    Circle.menu.data.name = name_obj
                    Circle.menu.setListName = self.Callbackname(child,Circle)

                    self.SoursesObjectsDict[child] = Atributs(type_link="object",type_object=tp ,graphic_item=Circle)  
                    self.SoursesObjectsChildren[parent].add(child)
                
                elif tp == "conductor":
                    child = TreeWidgetItem(["",name_obj])
                    child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
                    child.setCheckState(0, Qt.Unchecked)
                    child.setIcon(1,QIcon("images/conductor.png"))
                    parent.addChild(child)

                    Polyline = GraphicsPolylineItem([[xy[0], -xy[1]] for xy in cord],data=data,calc_func=self.PointCalc if self.pcalc else None)
                    Polyline.hndl*=1/scl_layers  
                    Polyline.updateHandlesPos()
                    Polyline.dragable(self.drag_state)
                    Polyline.menu.data.name = name_obj
                    Polyline.menu.setListName = self.Callbackname(child,Polyline)

                    self.SoursesObjectsDict[child] = Atributs(type_link="object",type_object=tp ,graphic_item=Polyline)  
                    self.SoursesObjectsChildren[parent].add(child)
        
        
        for name_lv, obj in Areas:
            parent = TreeWidgetItem(self.AreasObjectsTree, ["",name_lv])
            parent.setFlags(parent.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsEditable)
            parent.setCheckState(0, Qt.Unchecked)
            parent.setIcon(1,QIcon("images/area.png"))

            self.AreasObjectsDict[parent] = Atributs()
            self.AreasObjectsChildren[parent] = set()

            for name_obj, tp, cord, data in obj:
                if tp == "horizontal_area":
                    child = TreeWidgetItem(["",name_obj])
                    child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
                    child.setCheckState(0, Qt.Unchecked)
                    child.setIcon(1,QIcon("images/rectangle.png"))
                    parent.addChild(child)

                    CalcArea = GraphicsRectItem((cord[0],-cord[3],cord[2],-cord[1]),data=data)
                    CalcArea.menu.data.name = name_obj
                    CalcArea.hndl*=1/scl_layers 
                    CalcArea.updateHandlesPos()
                    CalcArea.dragable(self.drag_state)
                    CalcArea.menu.setListName = self.Callbackname(child,CalcArea)

                    self.AreasObjectsDict[child] = Atributs(type_link="object",type_object=tp,graphic_item=CalcArea)    
                    self.AreasObjectsChildren[parent].add(child)

                elif tp == "vertical_area":
                    child = TreeWidgetItem(["",name_obj])
                    child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
                    child.setCheckState(0, Qt.Unchecked)
                    child.setIcon(1,QIcon("images/line.png"))
                    parent.addChild(child)

                    CalcArea = GraphicsLineItem((cord[0],-cord[1],cord[2],-cord[3]),data=data)
                    CalcArea.menu.data.name = name_obj   
                    CalcArea.hndl*=1/scl_layers 
                    CalcArea.updateHandlesPos()
                    CalcArea.dragable(self.drag_state)
                    CalcArea.menu.setListName = self.Callbackname(child,CalcArea)

                    self.AreasObjectsDict[child] = Atributs(type_link="object",type_object="vertical_area",graphic_item=CalcArea)
                    self.AreasObjectsChildren[parent].add(child)

                elif tp == "one_point":
                    child = TreeWidgetItem(["",name_obj])
                    child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
                    child.setCheckState(0, Qt.Unchecked)
                    child.setIcon(1,QIcon("images/point.png"))
                    parent.addChild(child)
                    
                    CalcArea = OneCalcCircle((cord[0],-cord[1]),data=data,calc_func=self.PointCalc if self.pcalc else None)
                    CalcArea.menu.data.name = name_obj
                    CalcArea.hndl*=1/scl_layers
                    CalcArea.updateHandlesPos()
                    CalcArea.dragable(self.drag_state)
                    CalcArea.menu.setListName = self.Callbackname(child,CalcArea)

                    self.AreasObjectsDict[child] = Atributs(type_link="object",type_object="one_point",graphic_item=CalcArea)
                    self.AreasObjectsChildren[parent].add(child)

        try:
            for name_lv, obj in Receivers:
                parent = TreeWidgetItem(self.ReceiverObjectsTree, ["",name_lv])
                parent.setFlags(parent.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsEditable)
                parent.setCheckState(0, Qt.Unchecked)
                parent.setIcon(1,QIcon("images/receiver.png"))

                self.ReceiverObjectsDict[parent] = Atributs()
                self.ReceiverObjectsChildren[parent] = set()

                for name_obj, tp, cord, data in obj:
                    if tp == "recconductor":
                        child = TreeWidgetItem(["",name_obj])
                        child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
                        child.setCheckState(0, Qt.Unchecked)
                        child.setIcon(1,QIcon("images/conductor.png"))
                        parent.addChild(child)

                        Polyline = RecPolylineItem([[xy[0], -xy[1]] for xy in cord],data=data)#,calc_func=self.PointCalc
                        Polyline.hndl*=1/scl_layers  
                        Polyline.updateHandlesPos()
                        Polyline.dragable(self.drag_state)
                        Polyline.menu.data.name = name_obj
                        Polyline.menu.setListName = self.Callbackname(child,Polyline)

                        self.ReceiverObjectsDict[child] = Atributs(type_link="object",type_object=tp ,graphic_item=Polyline)  
                        self.ReceiverObjectsChildren[parent].add(child)

        except Exception as ex:
            print(ex, "fdfdfdfd")


    def SaveCalcData(self):
        try:
            fname = QFileDialog.getSaveFileName(self, 'Сохранить файл', self.path_home,'*.mfc')
            if fname[0] == "" and fname[1] == "": return
            fname = fname[0]

            Sourses = []
            Areas = []
            Receivers = []

            for key,item in self.SoursesObjectsChildren.items():
                Sourses.append([key.text(1),[self.SoursesObjectsDict[i].graphic_item.menu.data.save() for i in item]]) 

            for key,item in self.AreasObjectsChildren.items():
                Areas.append([key.text(1),[self.AreasObjectsDict[i].graphic_item.menu.data.save() for i in item]])

            for key,item in self.ReceiverObjectsChildren.items():
                Receivers.append([key.text(1),[self.ReceiverObjectsDict[i].graphic_item.menu.data.save() for i in item]])

            settings = {"da":self.da.value(), "dl":self.dl.value()}
  
            with open( fname, "w", encoding="utf8") as f:
                json.dump({"Sourses":Sourses, "Areas":Areas, "Receivers":Receivers, "Settings":settings},f, indent=4)
        except Exception as ex:
            print(ex)
        else:
            print("good save")
           
        
if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = Screen()
    ex.show()
    sys.exit(app.exec_())